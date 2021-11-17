import openpyxl
from openpyxl.styles import Border, Side, colors
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

"""合并单元格"""
result_file = 'test.xlsx'
wb = openpyxl.load_workbook(result_file)
ws = wb['Sheet1']

# 获取第一列数据
type_list = []
i = 2
while True:
    r = ws.cell(i, 2).value
    if r:
        type_list.append(r)
    else:
        break
    i += 1

# 判断合并单元格的始末位置
s = 0
e = 0
on = 1
flag = type_list[0]
#设置数字格式，
for i in range(len(type_list)):
    ws.cell(i + 2, 6).number_format = '0.00'
    ws.cell(i + 2, 7).number_format = '0.000'
    ws.cell(i + 2, 8).number_format = '0.000'
    ws.cell(i + 2, 7).alignment = Alignment(horizontal='center', vertical='center',
                                            text_rotation=0, wrap_text=True, shrink_to_fit=False,
                                            indent=0)
    ws.cell(i + 2, 8).alignment = Alignment(horizontal='center', vertical='center',
                                            text_rotation=0, wrap_text=True, shrink_to_fit=False,
                                            indent=0)

# 边框设置
border_set = Border(left=Side(style='thin', color=colors.BLACK),
                    right=Side(style='thin', color=colors.BLACK),
                    top=Side(style='thin', color=colors.BLACK),
                    bottom=Side(style='thin', color=colors.BLACK))
# 设置字体
font1 = Font(name=u'Arial', bold=False, italic=False, size=10.5)
font2 = Font(name=u'宋体', bold=True, italic=False, size=10.5)
#设置前景色
fill = PatternFill(fill_type = 'solid',start_color= 'BFBFBF',end_color='BFBFBF')


#装置边框，字体，前景色
for i in range(0, len(type_list)+1):
    for j in range(1, 9):
        ws.cell(row=1 + i, column=j).border = border_set  # 设置单元格格式
        ws.cell(row=1 + i, column=j).font = font1 #设置字体
        ws.cell(row=1,column=j).font = font2
        ws.cell(1,column = j).fill = fill #设置前景色

ws.column_dimensions['C'].width = 12.5
ws.column_dimensions['D'].width = 11.5
ws.column_dimensions['G'].width = 13.5
ws.column_dimensions['H'].width = 13.5
for i in range(len(type_list)):
    if type_list[i] != flag:
        flag = type_list[i]
        e = i - 1
        for a in range(1, 7):
            ws.cell(s + 2, a).alignment = Alignment(horizontal='center', vertical='center',
                                                    text_rotation=0, wrap_text=True, shrink_to_fit=False,
                                                    indent=0)

        if e >= s:
            ws['A' + str(s + 2)] = on
            ws.merge_cells('B' + str(s + 2) + ':B' + str(e + 2))
            ws.merge_cells('C' + str(s + 2) + ':C' + str(e + 2))
            ws.merge_cells('D' + str(s + 2) + ':D' + str(e + 2))
            ws.merge_cells('E' + str(s + 2) + ':E' + str(e + 2))
            ws.merge_cells('F' + str(s + 2) + ':F' + str(e + 2))
            ws.merge_cells('A' + str(s + 2) + ':A' + str(e + 2))
            s = e + 1
            on = on + 1


            # ws.cell(s + 2, 3).alignment = Alignment(horizontal='center', vertical='center',
            #                                         text_rotation=0, wrap_text=True, shrink_to_fit=False,
            #                                         indent=0)
    if i == len(type_list) - 1:
        ws['A' + str(s + 2)] = on
        e = i
        ws.merge_cells('B' + str(s + 2) + ':B' + str(e + 2))
        ws.merge_cells('C' + str(s + 2) + ':C' + str(e + 2))
        ws.merge_cells('D' + str(s + 2) + ':D' + str(e + 2))
        ws.merge_cells('E' + str(s + 2) + ':E' + str(e + 2))
        ws.merge_cells('F' + str(s + 2) + ':F' + str(e + 2))
        ws.merge_cells('A' + str(s + 2) + ':A' + str(e + 2))
        ws.cell(s + 2, 6).number_format = '0.00'
        ws.cell(s + 2, 7).number_format = '0.000'
        ws.cell(s + 2, 8).number_format = '0.000'
        for a in range(1, 7):
            ws.cell(s + 2, a).alignment = Alignment(horizontal='center', vertical='center',
                                                    text_rotation=0, wrap_text=True, shrink_to_fit=False,
                                                    indent=0)



wb.save("实验农房test.xlsx")
